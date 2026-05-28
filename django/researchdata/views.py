from django.views.generic import TemplateView
from openpyxl.utils import get_column_letter
from django.apps import apps
from django.http import HttpResponse, HttpResponseBadRequest
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from . import models
import csv
import io
import openpyxl


class DownloadsHomeTemplateView(LoginRequiredMixin, TemplateView):
    """
    Class-based view to show the Download home template
    """
    template_name = 'researchdata/downloads-home.html'


@login_required
def downloads_excel_templates(request):
    """
    A view to download an Excel spreadsheet that includes a tab
    for each specified model. Each tab includes a header row,
    which includes all field names.
    This is used for the team to fill out if the system goes down.
    """

    # 1. Define your list of models
    models_list = [
        models.SurveyRecord,
        models.SurveyUnitMaterialsCountedAndCollected,

        models.Feature,

        models.GriddedCollection,
        models.GridSquare,

        models.BulkMaterial,
        models.BulkMaterialBatch,
        models.FlaggedItem,

        models.SpecialistStudy
    ]

    # 2. Initialize a new Excel Workbook in memory
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # 3. Loop through each model to create its own tab
    for model in models_list:
        # Create a new sheet named after the model
        # Note: Excel sheet names are strictly limited to 31 characters
        sheet_title = model.__name__[:31]
        sheet = wb.create_sheet(title=sheet_title)

        # Extract standard fields (Char, Integer, ForeignKey)
        # AND ManyToMany fields (ignoring reverse relations)
        standard_fields = model._meta.fields
        m2m_fields = model._meta.many_to_many

        # Get the string names for all gathered fields
        header_row = [field.name for field in standard_fields] + \
                     [field.name for field in m2m_fields]

        # 4. Write the header row to the top of the sheet
        sheet.append(header_row)

        # 5. Auto-fit columns
        # enumerate(..., start=1) is used because openpyxl column indexes start at 1, not 0
        for col_idx, column_title in enumerate(header_row, start=1):
            col_letter = get_column_letter(col_idx)
            # Calculate width based on string length, plus a little padding so it isn't cramped
            adjusted_width = len(str(column_title)) + 4
            # Apply the width to the column
            sheet.column_dimensions[col_letter].width = adjusted_width

    # 6. Save the workbook to an in-memory buffer instead of the hard drive
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)  # Reset the buffer's "cursor" to the beginning

    # 7. Construct the HTTP response to trigger a file download
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="artemis_database_template.xlsx"'

    return response


@login_required
def download_model_data_csv(request):
    """
    A view to return CSV data for the given model
    Model name to be provided in get request, e.g. <url>?model=SurveyRecord
    """

    # 1. Get the model name from the URL parameter (e.g., ?model=Product)
    model_name = request.GET.get('model')

    if not model_name:
        return HttpResponseBadRequest("Missing 'model' parameter in the URL.")

    # 2. Find the actual Model class by searching installed apps
    target_model = None
    for app_model in apps.get_models():
        if app_model.__name__.lower() == model_name.lower():
            target_model = app_model
            break

    if not target_model:
        return HttpResponseBadRequest(f"Model '{model_name}' not found.")

    # 3. Set up the CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{target_model.__name__.lower()}_data.csv"'
    writer = csv.writer(response)

    # 4. Extract field references
    standard_fields = target_model._meta.fields
    m2m_fields = target_model._meta.many_to_many

    # 5. Write the header row
    header_row = [field.name for field in standard_fields]\
        + [field.name for field in m2m_fields]
    writer.writerow(header_row)

    # 6. Fetch the data (using prefetch_related for M2M performance)
    m2m_field_names = [field.name for field in m2m_fields]
    queryset = target_model.objects.all()

    if m2m_field_names:
        queryset = queryset.prefetch_related(*m2m_field_names)

    # 7. Iterate through the database records and write the rows
    for instance in queryset:
        row = []

        # Grab data for standard columns
        for field in standard_fields:
            # Using getattr retrieves the actual value for this specific field
            value = getattr(instance, field.name)
            row.append(value)

        # Grab and format data for ManyToMany columns
        for field in m2m_fields:
            m2m_manager = getattr(instance, field.name)

            # Convert each related object to a string and join them with semicolons
            m2m_strings = [str(related_obj) for related_obj in m2m_manager.all()]
            joined_string = "; ".join(m2m_strings)

            row.append(joined_string)

        writer.writerow(row)

    return response
